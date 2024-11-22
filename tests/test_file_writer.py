import os
from zipfile import ZipFile
from openpyxl import load_workbook
from src.file_writer import write_zip, write_excel

def test_write_excel(tmp_path):
    """Test an Excel file with multiple sheets."""
    data1 = [
        {
            "team_name": "Team A",
            "year": "1990",
            "wins": 10,
            "losses": 5,
            "ot_losses": 3,
            "win_pct": 0.667,
            "goals_for": 100,
            "goals_against": 90,
            "goal_difference": 10,
        },
        {
            "team_name": "Team B",
            "year": "1990",
            "wins": 8,
            "losses": 7,
            "ot_losses": 3,
            "win_pct": 0.533,
            "goals_for": 80,
            "goals_against": 85,
            "goal_difference": -5,
        },
    ]
    data2 = [
        {"year": "1990", "winner": "Team A", "winner_wins": 10, "loser": "Team B", "loser_wins": 8}
    ]
    excel_file = tmp_path / "output.xlsx"

    # Write Excel file
    write_excel(data1, data2, excel_file)

    # Verify Excel file exists
    assert excel_file.exists()

    # Load workbook and verify content
    workbook = load_workbook(excel_file)

    # Check Sheet 1: All Data
    sheet1 = workbook["NHL Stats 1990-2011"]
    assert sheet1.title == "NHL Stats 1990-2011"
    assert sheet1.max_row == 3  # Header + 2 rows
    assert sheet1.max_column == 9  # Number of columns

    # Validate Sheet 1 Content
    assert sheet1.cell(1, 1).value == "Team Name"  # Header
    assert sheet1.cell(2, 1).value == "Team A"    # Row 1, Column 1
    assert sheet1.cell(2, 3).value == 10          # Row 1, Column 3 (Wins)
    assert sheet1.cell(3, 3).value == 8           # Row 2, Column 3 (Wins)

    # Check Sheet 2: Winner and Loser per Year
    sheet2 = workbook["Winner and Loser per Year"]
    assert sheet2.title == "Winner and Loser per Year"
    assert sheet2.max_row == 2  # Header + 1 row
    assert sheet2.max_column == 5  # Number of columns

    # Validate Sheet 2 Content
    assert sheet2.cell(1, 1).value == "Year"      # Header
    assert sheet2.cell(2, 2).value == "Team A"    # Winner
    assert sheet2.cell(2, 4).value == "Team B"    # Loser

def test_write_excel_empty_data(tmp_path):
    """Test writing Excel file with empty data."""
    data1 = []  # No data for Sheet 1
    data2 = []  # No data for Sheet 2
    excel_file = tmp_path / "empty_output.xlsx"

    # Write Excel file
    write_excel(data1, data2, excel_file)

    # Verify Excel file exists
    assert excel_file.exists()

    # Load workbook
    workbook = load_workbook(excel_file)

    # Check Sheet 1
    sheet1 = workbook["NHL Stats 1990-2011"]
    assert sheet1.max_row == 1  # Only header
    assert sheet1.max_column == 9  # Number of columns

    # Check Sheet 2
    sheet2 = workbook["Winner and Loser per Year"]
    assert sheet2.max_row == 1  # Only header
    assert sheet2.max_column == 5  # Number of columns
